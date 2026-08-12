"""
GOST Report Template Engine — Sprint 6.

Pre-configured report templates for Russian regulatory reporting:
- Бухгалтерский баланс (форма 0710001)
- Отчёт о финансовых результатах (форма 0710002)
- 6-НДФЛ
- СЗВ-М
- Статистические формы Росстата

Renders to PDF (WeasyPrint), Excel (openpyxl), and XML (ФНС format).
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any

logger = logging.getLogger("gost_bi.gost.templates")


class GOSTReportFormat(str, Enum):
    PDF = "pdf"
    EXCEL = "xlsx"
    XML_FNS = "xml_fns"
    CSV_1251 = "csv_1251"


@dataclass
class GOSTTemplate:
    name: str
    code: str
    description: str
    category: str
    version: str = "1.0"
    fields: list[dict[str, Any]] = field(default_factory=list)
    valid_from: str = ""
    valid_to: str = ""


@dataclass
class GOSTReport:
    template: GOSTTemplate
    data: dict[str, Any]
    generated_at: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    rendered: bytes = b""


BUILTIN_TEMPLATES: list[GOSTTemplate] = [
    GOSTTemplate(
        name="Бухгалтерский баланс",
        code="0710001",
        description="Форма бухгалтерского баланса (ОКУД 0710001)",
        category="Бухгалтерская отчётность",
        fields=[
            {"code": "1110", "name": "Нематериальные активы", "section": "АКТИВ", "line": 1},
            {"code": "1150", "name": "Основные средства", "section": "АКТИВ", "line": 2},
            {"code": "1210", "name": "Запасы", "section": "АКТИВ", "line": 3},
            {"code": "1230", "name": "Дебиторская задолженность", "section": "АКТИВ", "line": 4},
            {"code": "1250", "name": "Денежные средства", "section": "АКТИВ", "line": 5},
            {"code": "1600", "name": "БАЛАНС (актив)", "section": "АКТИВ", "line": 6},
            {"code": "1310", "name": "Уставный капитал", "section": "ПАССИВ", "line": 7},
            {"code": "1370", "name": "Нераспределённая прибыль", "section": "ПАССИВ", "line": 8},
            {"code": "1410", "name": "Долгосрочные заёмные средства", "section": "ПАССИВ", "line": 9},
            {"code": "1510", "name": "Краткосрочные заёмные средства", "section": "ПАССИВ", "line": 10},
            {"code": "1520", "name": "Кредиторская задолженность", "section": "ПАССИВ", "line": 11},
            {"code": "1700", "name": "БАЛАНС (пассив)", "section": "ПАССИВ", "line": 12},
        ],
    ),
    GOSTTemplate(
        name="Отчёт о финансовых результатах",
        code="0710002",
        description="Отчёт о прибылях и убытках (ОКУД 0710002)",
        category="Бухгалтерская отчётность",
        fields=[
            {"code": "2110", "name": "Выручка", "line": 1},
            {"code": "2120", "name": "Себестоимость продаж", "line": 2},
            {"code": "2100", "name": "Валовая прибыль (убыток)", "line": 3},
            {"code": "2210", "name": "Коммерческие расходы", "line": 4},
            {"code": "2220", "name": "Управленческие расходы", "line": 5},
            {"code": "2200", "name": "Прибыль (убыток) от продаж", "line": 6},
            {"code": "2300", "name": "Прибыль до налогообложения", "line": 7},
            {"code": "2410", "name": "Налог на прибыль", "line": 8},
            {"code": "2400", "name": "Чистая прибыль (убыток)", "line": 9},
        ],
    ),
    GOSTTemplate(
        name="6-НДФЛ",
        code="6NDFL",
        description="Расчёт сумм налога на доходы физических лиц",
        category="Налоговая отчётность",
        fields=[
            {"code": "010", "name": "Ставка налога", "line": 1},
            {"code": "020", "name": "Сумма дохода", "line": 2},
            {"code": "030", "name": "Сумма вычетов", "line": 3},
            {"code": "040", "name": "Исчисленная сумма налога", "line": 4},
            {"code": "070", "name": "Удержанная сумма налога", "line": 5},
        ],
    ),
    GOSTTemplate(
        name="СЗВ-М",
        code="SZV-M",
        description="Сведения о застрахованных лицах (ежемесячная)",
        category="Пенсионный фонд",
        fields=[
            {"code": "FIO", "name": "Фамилия, имя, отчество", "line": 1},
            {"code": "SNILS", "name": "СНИЛС", "line": 2},
            {"code": "INN", "name": "ИНН", "line": 3},
        ],
    ),
    GOSTTemplate(
        name="СЗВ-ТД",
        code="SZV-TD",
        description="Сведения о трудовой деятельности зарегистрированного лица",
        category="Пенсионный фонд",
        fields=[
            {"code": "FIO", "name": "Фамилия, имя, отчество", "line": 1},
            {"code": "SNILS", "name": "СНИЛС", "line": 2},
            {"code": "EVENT_DATE", "name": "Дата кадрового мероприятия", "line": 3},
            {"code": "EVENT_TYPE", "name": "Вид мероприятия (ПРИЁМ, ПЕРЕВОД, УВОЛЬНЕНИЕ)", "line": 4},
            {"code": "POSITION", "name": "Должность", "line": 5},
            {"code": "DEPARTMENT", "name": "Подразделение", "line": 6},
        ],
    ),
    GOSTTemplate(
        name="2-НДФЛ",
        code="2NDFL",
        description="Справка о доходах и суммах налога физического лица",
        category="Налоговая отчётность",
        fields=[
            {"code": "INN", "name": "ИНН", "line": 1},
            {"code": "FIO", "name": "Фамилия, имя, отчество", "line": 2},
            {"code": "MONTH", "name": "Месяц", "line": 3},
            {"code": "INCOME_CODE", "name": "Код дохода", "line": 4},
            {"code": "INCOME_AMOUNT", "name": "Сумма дохода", "line": 5},
            {"code": "DEDUCTION_CODE", "name": "Код вычета", "line": 6},
            {"code": "DEDUCTION_AMOUNT", "name": "Сумма вычета", "line": 7},
            {"code": "TAX_AMOUNT", "name": "Сумма налога", "line": 8},
        ],
    ),
    GOSTTemplate(
        name="П-4",
        code="P-4",
        description="Сведения о численности и заработной плате работников (Росстат)",
        category="Статистика (Росстат)",
        fields=[
            {"code": "HEADCOUNT", "name": "Среднесписочная численность", "line": 1},
            {"code": "PAYROLL_FUND", "name": "Фонд начисленной заработной платы", "line": 2},
            {"code": "SOCIAL_PAYMENTS", "name": "Выплаты социального характера", "line": 3},
            {"code": "HOURS_WORKED", "name": "Отработано человеко-часов", "line": 4},
        ],
    ),
]


class GOSTReportRenderer:
    """Renders GOST reports to various formats."""

    @staticmethod
    def render_pdf(report: GOSTReport) -> bytes:
        try:
            from weasyprint import HTML

            html = GOSTReportRenderer._build_html(report)
            return HTML(string=html).write_pdf()
        except ImportError:
            logger.warning("WeasyPrint not installed — PDF rendering skipped")
            return f"PDF placeholder for {report.template.name}".encode()

    @staticmethod
    def render_excel(report: GOSTReport) -> bytes:
        try:
            from io import BytesIO

            import openpyxl
            from openpyxl.styles import Border, Font, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report.template.name[:31]

            ws["A1"] = report.template.name
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:D1")

            ws["A2"] = f"Форма {report.template.code}"
            ws["A3"] = f"Сформировано: {report.generated_at}"

            headers = ["Код", "Наименование", "Значение", "Раздел"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)

            for i, field in enumerate(report.template.fields, 6):
                code = field.get("code", "")
                name = field.get("name", "")
                value = report.data.get(code, 0)
                section = field.get("section", "")

                ws.cell(row=i, column=1, value=code)
                ws.cell(row=i, column=2, value=name)
                ws.cell(row=i, column=3, value=value)
                ws.cell(row=i, column=4, value=section)

            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()

        except ImportError:
            logger.warning("openpyxl not installed — Excel rendering skipped")
            return f"Excel placeholder for {report.template.name}".encode()

    @staticmethod
    def render_xml_fns(report: GOSTReport) -> bytes:
        from xml.etree import ElementTree as ET

        root = ET.Element("Файл", {"ИдФайл": f"NO_{report.template.code}_{report.generated_at[:10].replace('-', '')}"})
        ET.SubElement(root, "Форма", {"КНД": report.template.code, "Дата": report.generated_at[:10]})

        for field in report.template.fields:
            code = field["code"]
            value = report.data.get(code, "")
            ET.SubElement(root, "Показатель", {"Код": code}).text = str(value)

        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    @staticmethod
    def _build_html(report: GOSTReport) -> str:
        rows = ""
        for field in report.template.fields:
            value = report.data.get(field["code"], "—")
            rows += f"<tr><td>{field['code']}</td><td>{field['name']}</td><td>{value}</td><td>{field.get('section', '')}</td></tr>"

        return f"""<!DOCTYPE html>
<html lang="ru">
<head><meta charset="utf-8"><title>{report.template.name}</title>
<style>
body {{ font-family: 'Arial', sans-serif; margin: 40px; }}
h1 {{ font-size: 16pt; text-align: center; }}
h2 {{ font-size: 12pt; text-align: center; color: #555; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
th, td {{ border: 1px solid #000; padding: 6px 10px; font-size: 10pt; }}
th {{ background: #f0f0f0; }}
</style></head>
<body>
<h1>{report.template.name}</h1>
<h2>Форма {report.template.code} | Сформировано: {report.generated_at[:10]}</h2>
<table>
<tr><th>Код</th><th>Наименование</th><th>Значение</th><th>Раздел</th></tr>
{rows}
</table>
</body></html>"""
